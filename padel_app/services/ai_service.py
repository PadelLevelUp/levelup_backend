"""
AI-powered Excel import analysis service.

Pipeline (dependency-ordered):
    1. parse_excel              -- parse workbook into sheet -> raw rows
    2. pick_relevant_sheets     -- cheap LLM call to filter irrelevant sheets
    3. detect_table_segments    -- find multiple tables within a single sheet
    4. column-map all segments  -- LLM maps columns to target schemas (~4s each)
    5. ordered extraction chain:
        a. Coach Levels     -- map Excel levels to coach's existing DB levels via LLM
        b. Players          -- validate level_code against mapped Coach Levels
        c. Eval Categories  -- validate names are real categories (not numbers)
        d. Classes          -- require date+times, auto-generate title if missing
        e. Players in Classes -- require known Player + known Class
        f. Presences        -- require known Player + status + date
        g. Evaluations      -- require known Player + known Category
        h. Strengths        -- require known Player
        i. Weaknesses       -- require known Player
    6. Cross-table validation   -- drop orphaned references, count drops
    7. LLM cleanup pass         -- only for tables with anomalies
    8. stream_import_analysis   -- orchestrate and yield SSE events

Design principles:
    - User selects which tables to import (default: all).
      Coach Levels and Eval Categories are auto-included as dependencies.
    - Column-mapping strategy for ALL sheets (fast, ~4s, reliable).
    - Business rules enforced programmatically, not by LLM.
    - LLM used only for: sheet relevance, column mapping, level mapping,
      name validation, and anomaly correction.
    - response_format=json_object on all LLM calls.
    - No pandas dependency.

All OpenAI calls use the openai >= 1.0 client API.
"""
from __future__ import annotations

import json
import logging
import os
import re
import time
from typing import Generator

from openai import OpenAI
import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

_client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY", ""))

_MODEL = "gpt-4o-mini"
_logger = logging.getLogger(__name__)

if not _logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    _logger.addHandler(_handler)
    _logger.setLevel(logging.INFO)

_MAX_RETRIES = 2
_RETRY_DELAY_S = 1.0

# All user-selectable tables
ALL_IMPORTABLE_TABLES: list[str] = [
    "Players",
    "Classes",
    "Players in Classes",
    "Presences",
    "Evaluations",
    "Strengths",
    "Weaknesses",
]

# Tables that are always included as dependencies (never user-deselected)
_AUTO_TABLES: set[str] = {"Coach Levels", "Evaluation Categories"}

_DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_TIME_PATTERN = re.compile(r"^\d{2}:\d{2}$")


def _elapsed_ms(start: float) -> float:
    return (time.perf_counter() - start) * 1000


def _log_timing(step: str, start: float, **meta) -> float:
    elapsed_ms = _elapsed_ms(start)
    if meta:
        meta_str = ", ".join(f"{k}={v}" for k, v in meta.items())
        _logger.info("[AI TIMER] %s took %.2fms (%s)", step, elapsed_ms, meta_str)
    else:
        _logger.info("[AI TIMER] %s took %.2fms", step, elapsed_ms)
    return elapsed_ms


# ---------------------------------------------------------------------------
# LLM helper
# ---------------------------------------------------------------------------

def _call_openai(
    *,
    messages: list[dict],
    max_tokens: int = 4096,
    temperature: float = 0,
    use_json_mode: bool = False,
    step_label: str = "openai_call",
    **extra_meta,
) -> str:
    """Call OpenAI with retry. json_mode guarantees valid JSON output."""
    kwargs: dict = {
        "model": _MODEL,
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": temperature,
    }
    if use_json_mode:
        kwargs["response_format"] = {"type": "json_object"}

    last_exc: Exception | None = None
    for attempt in range(_MAX_RETRIES + 1):
        try:
            llm_start = time.perf_counter()
            response = _client.chat.completions.create(**kwargs)
            _log_timing(
                f"{step_label}.openai_call", llm_start,
                model=_MODEL, attempt=attempt, **extra_meta,
            )
            return response.choices[0].message.content.strip()
        except Exception as exc:
            last_exc = exc
            _logger.warning("[AI] %s attempt %d failed: %s", step_label, attempt, exc)
            if attempt < _MAX_RETRIES:
                time.sleep(_RETRY_DELAY_S * (attempt + 1))
    raise last_exc  # type: ignore[misc]


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def _is_empty(val) -> bool:
    return val is None or (isinstance(val, str) and val.strip() == "")


def _is_numeric(val) -> bool:
    if isinstance(val, (int, float)):
        return True
    if val is None:
        return False
    try:
        float(str(val))
        return True
    except (ValueError, TypeError):
        return False


def _deduplicate_rows(existing: list[dict], new_rows: list[dict]) -> list[dict]:
    seen: dict[str, dict] = {}
    for row in existing + new_rows:
        key = json.dumps(row, default=str, sort_keys=True)
        seen[key] = row
    return list(seen.values())


def _merge_into(analysis: dict[str, list], new_tables: dict[str, list]) -> None:
    for table_name, table_rows in new_tables.items():
        if isinstance(table_rows, list) and table_rows:
            analysis[table_name] = _deduplicate_rows(
                analysis.get(table_name, []), table_rows
            )


# ---------------------------------------------------------------------------
# Step 1 -- Excel parsing
# ---------------------------------------------------------------------------

def parse_excel(file_bytes: bytes) -> dict[str, list[list]]:
    """Parse workbook into {sheet_name: raw_rows} (lists of cell values).
    Keeps raw rows so segment detection can find embedded headers."""
    from io import BytesIO

    total_start = time.perf_counter()
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    _log_timing("parse_excel.load_workbook", total_start, sheets=len(wb.sheetnames))
    sheets: dict[str, list[list]] = {}

    for sheet_name in wb.sheetnames:
        sheet_start = time.perf_counter()
        ws = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if len(rows) < 2:
            _log_timing("parse_excel.sheet", sheet_start,
                        sheet=sheet_name, status="skipped_short", rows=len(rows))
            continue
        has_data = any(
            any(c is not None and c != "" for c in row)
            for row in rows[1:]
        )
        if has_data:
            sheets[sheet_name] = rows
            _log_timing("parse_excel.sheet", sheet_start,
                        sheet=sheet_name, status="included", rows=len(rows))
        else:
            _log_timing("parse_excel.sheet", sheet_start,
                        sheet=sheet_name, status="empty", rows=0)

    wb.close()
    _log_timing("parse_excel.total", total_start, selected_sheets=len(sheets))
    return sheets


# ---------------------------------------------------------------------------
# Step 2 -- Sheet relevance filter
# ---------------------------------------------------------------------------

def pick_relevant_sheets(
    sheets_data: dict[str, list[list]],
    requested_tables: set[str],
) -> list[str]:
    """Use a cheap LLM call to decide which sheets are relevant,
    scoped to only the tables the user wants to import."""
    total_start = time.perf_counter()
    if not sheets_data:
        return []

    tables_desc = ", ".join(sorted(requested_tables | _AUTO_TABLES))

    summaries = []
    for name, rows in sheets_data.items():
        header = [str(h)[:60] if h is not None else "" for h in rows[0]]
        sample = [
            [str(c)[:60] if c is not None else None for c in row]
            for row in rows[1:3]
        ]
        summaries.append({"sheet": name, "headers": header, "sample_rows": sample})

    prompt = (
        "You are a data import assistant for a padel coaching app.\n\n"
        f"The user wants to import ONLY these tables: {tables_desc}.\n\n"
        "Sheet names may be in ANY language. "
        "Judge by column headers and data content, NOT by name.\n\n"
        'Return a JSON object: {"sheets": ["name1", "name2", ...]} '
        "with sheet names that contain data relevant to the requested tables. "
        "When in doubt, include the sheet.\n\n"
        f"Sheets:\n{json.dumps(summaries, default=str, indent=2)}"
    )

    try:
        content = _call_openai(
            messages=[{"role": "user", "content": prompt}],
            max_tokens=256, use_json_mode=True,
            step_label="pick_relevant_sheets",
        )
        result = json.loads(content)
        selected = result.get("sheets", result) if isinstance(result, dict) else result
        _log_timing("pick_relevant_sheets.total", total_start,
                    selected=len(selected), total=len(sheets_data))
        return selected
    except Exception:
        _logger.exception("[AI] pick_relevant_sheets failed, returning all")
        return list(sheets_data.keys())


# ---------------------------------------------------------------------------
# Step 3 -- Table segment detection
# ---------------------------------------------------------------------------

class TableSegment:
    """A contiguous block of rows within a sheet that forms one logical table."""
    __slots__ = ("sheet_name", "segment_index", "headers", "data_rows", "start_row")

    def __init__(self, sheet_name: str, segment_index: int,
                 headers: list[str], data_rows: list[dict], start_row: int):
        self.sheet_name = sheet_name
        self.segment_index = segment_index
        self.headers = headers
        self.data_rows = data_rows
        self.start_row = start_row

    @property
    def label(self) -> str:
        if self.segment_index == 0:
            return self.sheet_name
        return f"{self.sheet_name} (segment {self.segment_index + 1})"

    def __repr__(self) -> str:
        return f"<Segment {self.label}: {len(self.data_rows)} rows, {len(self.headers)} cols>"


def _is_likely_header(row: list, prev_row_types: list[type] | None) -> bool:
    if not row:
        return False
    non_null = [c for c in row if c is not None and c != ""]
    if len(non_null) < 2:
        return False
    all_str = all(isinstance(c, str) for c in non_null)
    if not all_str:
        return False
    if prev_row_types:
        numeric_before = sum(1 for t in prev_row_types if t in (int, float))
        if numeric_before >= len(prev_row_types) * 0.3:
            return True
    return False


def detect_table_segments(sheet_name: str, raw_rows: list[list]) -> list[TableSegment]:
    """Find one or more logical tables within a sheet."""
    total_start = time.perf_counter()
    if not raw_rows:
        return []

    header_idx = 0
    for i, row in enumerate(raw_rows):
        if any(c is not None and c != "" for c in row):
            header_idx = i
            break

    segments: list[TableSegment] = []
    current_header_idx = header_idx
    current_data_start = header_idx + 1
    consecutive_blank = 0
    prev_types: list[type] | None = None

    def _finalise_segment(data_end: int) -> None:
        nonlocal current_header_idx
        header_row = raw_rows[current_header_idx]
        headers = [
            str(h).strip() if h is not None else f"col_{j}"
            for j, h in enumerate(header_row)
        ]
        data_rows = []
        for r in raw_rows[current_data_start:data_end]:
            row_dict = {
                headers[j]: (r[j] if j < len(r) else None)
                for j in range(len(headers))
            }
            if not all(v is None or v == "" for v in row_dict.values()):
                data_rows.append(row_dict)
        if data_rows:
            segments.append(TableSegment(
                sheet_name=sheet_name,
                segment_index=len(segments),
                headers=headers,
                data_rows=data_rows,
                start_row=current_header_idx,
            ))

    i = current_data_start
    while i < len(raw_rows):
        row = raw_rows[i]
        is_blank = all(c is None or c == "" for c in row)

        if is_blank:
            consecutive_blank += 1
            if consecutive_blank >= 2:
                _finalise_segment(i - consecutive_blank + 1)
                for j in range(i + 1, len(raw_rows)):
                    if any(c is not None and c != "" for c in raw_rows[j]):
                        current_header_idx = j
                        current_data_start = j + 1
                        i = j + 1
                        consecutive_blank = 0
                        prev_types = None
                        break
                else:
                    i = len(raw_rows)
                continue
            i += 1
            continue

        consecutive_blank = 0
        if prev_types and _is_likely_header(row, prev_types):
            _finalise_segment(i)
            current_header_idx = i
            current_data_start = i + 1
            prev_types = None
            i += 1
            continue

        prev_types = [type(c) for c in row]
        i += 1

    _finalise_segment(len(raw_rows))
    _log_timing("detect_table_segments", total_start,
                sheet=sheet_name, segments=len(segments),
                total_rows=sum(len(s.data_rows) for s in segments))
    return segments


# ---------------------------------------------------------------------------
# Step 4 -- Column-mapping (LLM)
# ---------------------------------------------------------------------------

def _column_profile(rows: list[dict]) -> dict[str, list]:
    """Build {column: [up to 15 unique non-null sample values]}."""
    result: dict[str, list] = {}
    for header in (rows[0].keys() if rows else []):
        seen: list[str] = []
        for row in rows:
            v = row.get(header)
            if v is not None and str(v) not in seen:
                seen.append(str(v))
            if len(seen) >= 15:
                break
        result[header] = seen
    return result


def _infer_column_mapping(
    segment: TableSegment,
    requested_tables: set[str],
) -> dict:
    """Ask the LLM to map source columns to target schema fields.
    Only maps to tables the user actually wants."""
    total_start = time.perf_counter()
    profile = _column_profile(segment.data_rows)

    # Build the table descriptions only for requested + auto tables
    active_tables = requested_tables | _AUTO_TABLES
    table_fields = {
        "Coach Levels": "code, label, display_order",
        "Evaluation Categories": "name, scale_min, scale_max",
        "Players": "name, email, phone, level_code, side",
        "Classes": "title, type (academy/private), is_recurring, day (YYYY-MM-DD), start_time (HH:MM), end_time (HH:MM), max_players",
        "Players in Classes": "lesson_title, player_name",
        "Presences": "lesson_title, date (YYYY-MM-DD), player_name, status (present/absent), justification (justified/unjustified)",
        "Evaluations": "player_name, date (YYYY-MM-DD), category_name, score",
        "Strengths": "player_name, strengths",
        "Weaknesses": "player_name, weaknesses",
    }
    tables_section = "\n".join(
        f'- "{t}": {table_fields[t]}'
        for t in table_fields if t in active_tables
    )

    prompt = f"""You are a data import assistant for a padel coaching app.

I have a table segment from Excel sheet "{segment.sheet_name}" with {len(segment.data_rows)} rows.
Column names and up to 15 unique sample values per column:

{json.dumps(profile, default=str, indent=2)}

Map this to ONE OR MORE of the following tables.
Return a JSON object with key "mappings" containing an array of mapping objects.

STANDARD mapping (for flat tables):
- "table_name": one of the table names below
- "columns": object mapping target_field to source_column_name (or null)
- "value_mappings": object mapping target_field to {{original_value: normalised_value}}

PIVOT mapping (for evaluation/scoring sheets where columns ARE categories):
If this sheet has player names in one column and multiple columns that are
evaluation categories (e.g. "Technique", "Tactics", "Physical", etc.) with
numeric scores in the cells, use a PIVOT mapping:
- "table_name": "Evaluations"
- "pivot": true
- "player_column": "<column containing player names>"
- "date_column": "<column containing dates, or null>"
- "category_columns": ["<col1>", "<col2>", ...] — ONLY columns that are real
  evaluation category names with numeric score values. Do NOT include columns
  that contain player names, dates, levels, emails, or other non-score data.

Also when using pivot mapping, include a SEPARATE mapping for Evaluation Categories:
- "table_name": "Evaluation Categories"
- "extract_from_pivot": true
- "category_names": ["<same category column names>"]

ONLY map to these tables:
{tables_section}

If this segment contains level or category data, include mappings with
"extract_unique_from" naming the source column to extract unique values from.

Example for standard mapping:
{{"mappings": [
  {{"table_name": "Players", "columns": {{"name": "Nome", "email": "Email"}}, "value_mappings": {{}}}},
  {{"table_name": "Coach Levels", "extract_unique_from": "Nivel", "columns": {{"code": "Nivel", "label": "Nivel"}}, "value_mappings": {{}}}}
]}}

Example for pivot evaluation sheet:
{{"mappings": [
  {{"table_name": "Evaluations", "pivot": true, "player_column": "Aluno", "date_column": null, "category_columns": ["Tecnica", "Tatica", "Fisico", "Mental"]}},
  {{"table_name": "Evaluation Categories", "extract_from_pivot": true, "category_names": ["Tecnica", "Tatica", "Fisico", "Mental"]}}
]}}"""

    content = _call_openai(
        messages=[{"role": "user", "content": prompt}],
        max_tokens=1024, use_json_mode=True,
        step_label="_infer_column_mapping",
        segment=segment.label, rows=len(segment.data_rows),
    )
    result = json.loads(content)
    _log_timing("_infer_column_mapping.total", total_start,
                segment=segment.label,
                tables=len(result.get("mappings", [])))
    return result


def _apply_column_mapping(rows: list[dict], mapping: dict) -> tuple[str, list[dict]]:
    """Apply a single mapping dict to rows. Returns (table_name, processed_rows)."""
    table_name: str = mapping.get("table_name", "")
    columns: dict = mapping.get("columns", {})
    value_mappings: dict = mapping.get("value_mappings", {})

    extract_col = mapping.get("extract_unique_from")
    extract_pivot = mapping.get("extract_from_pivot")

    # Handle "extract_from_pivot" for Evaluation Categories from pivot sheets
    if extract_pivot:
        cat_names = mapping.get("category_names", [])
        processed = []
        for name in cat_names:
            name_str = str(name).strip()
            if name_str and not _is_numeric(name_str) and len(name_str) >= 2:
                processed.append({
                    "name": name_str,
                    "scale_min": 0,
                    "scale_max": 10,
                })
        return table_name, processed

    if extract_col:
        seen_values: set[str] = set()
        processed = []
        for row in rows:
            val = row.get(extract_col)
            if val is not None:
                val_str = str(val).strip()
                if val_str and val_str not in seen_values:
                    seen_values.add(val_str)
                    record: dict = {}
                    for field, source_col in columns.items():
                        if source_col:
                            record[field] = val_str
                    if table_name == "Coach Levels":
                        record.setdefault("display_order", len(processed) + 1)
                    processed.append(record)
        return table_name, processed

    processed = []
    for row in rows:
        record: dict = {}
        for field, source_col in columns.items():
            if not source_col:
                continue
            val = row.get(source_col)
            if hasattr(val, "strftime"):
                val = val.strftime("%Y-%m-%d")
            elif hasattr(val, "isoformat"):
                val = str(val)
            elif val is not None and not isinstance(val, (int, float, bool)):
                val = str(val)
            if field in value_mappings and val is not None:
                val = value_mappings[field].get(str(val), val)
            record[field] = val
        processed.append(record)

    return table_name, processed


def _process_segment(
    segment: TableSegment,
    requested_tables: set[str],
) -> dict[str, list[dict]]:
    """Map a segment via LLM then apply locally.

    Special handling: if the LLM identifies this as a pivot-style evaluation
    sheet (player rows x category columns with scores in cells), it returns
    a "pivot" mapping that we unpivot locally.
    """
    total_start = time.perf_counter()
    mapping_response = _infer_column_mapping(segment, requested_tables)
    mappings = mapping_response.get("mappings", [])
    if not mappings and "table_name" in mapping_response:
        mappings = [mapping_response]

    result: dict[str, list[dict]] = {}
    for mapping in mappings:
        # Handle pivot-style evaluation mapping
        if mapping.get("pivot"):
            tbl_name, processed = _apply_pivot_mapping(segment.data_rows, mapping)
        else:
            tbl_name, processed = _apply_column_mapping(segment.data_rows, mapping)
        if tbl_name and processed:
            result.setdefault(tbl_name, []).extend(processed)

    _log_timing("_process_segment.total", total_start,
                segment=segment.label, rows=len(segment.data_rows),
                tables=len(result),
                records=sum(len(v) for v in result.values()))
    return result


def _apply_pivot_mapping(
    rows: list[dict],
    mapping: dict,
) -> tuple[str, list[dict]]:
    """Unpivot a player-x-category evaluation sheet.

    The mapping should contain:
        - table_name: "Evaluations"
        - pivot: true
        - player_column: name of the column containing player names
        - date_column: name of the column containing dates (or null)
        - category_columns: list of column names that are evaluation categories
    """
    table_name = mapping.get("table_name", "Evaluations")
    player_col = mapping.get("player_column", "")
    date_col = mapping.get("date_column")
    category_cols = mapping.get("category_columns", [])

    if not player_col or not category_cols:
        return table_name, []

    processed = []
    for row in rows:
        player = row.get(player_col)
        if _is_empty(player):
            continue
        player_name = str(player).strip()

        date_val = None
        if date_col:
            d = row.get(date_col)
            if not _is_empty(d):
                if hasattr(d, "strftime"):
                    date_val = d.strftime("%Y-%m-%d")
                else:
                    date_str = str(d).strip()
                    m = re.match(r"(\d{4}-\d{2}-\d{2})", date_str)
                    date_val = m.group(1) if m else date_str

        for cat_col in category_cols:
            score = row.get(cat_col)
            if _is_empty(score) or not _is_numeric(score):
                continue
            processed.append({
                "player_name": player_name,
                "date": date_val,
                "category_name": str(cat_col).strip(),
                "score": float(score),
            })

    return table_name, processed


# ---------------------------------------------------------------------------
# Step 5 -- Ordered extraction & business validation
# ---------------------------------------------------------------------------

def _map_coach_levels(
    raw_levels: list[dict],
    raw_players: list[dict],
    existing_levels: list[dict],
) -> tuple[list[dict], dict[str, str]]:
    """Map raw Excel level values to coach's existing DB levels via LLM.

    Collects level values from both the raw Coach Levels table AND the
    level_code field in raw Players, since those are the actual values
    that need mapping.

    existing_levels: [{code, label, display_order}, ...] from coach_service.

    Returns:
        - mapped_levels: the existing DB levels that got matched (unchanged)
        - level_mapping: dict mapping raw Excel level string -> DB level code
          Only contains entries that map to a real DB level.
    """
    total_start = time.perf_counter()

    # Collect ALL unique raw level values from multiple sources
    raw_values: set[str] = set()
    for r in raw_levels:
        code = r.get("code")
        if not _is_empty(code):
            raw_values.add(str(code).strip())
        label = r.get("label")
        if not _is_empty(label):
            raw_values.add(str(label).strip())
    for p in raw_players:
        lc = p.get("level_code")
        if not _is_empty(lc):
            raw_values.add(str(lc).strip())

    if not raw_values:
        _log_timing("_map_coach_levels", total_start, status="no_raw_values")
        return [], {}

    if not existing_levels:
        # No existing levels in DB — cannot map, return empty
        # (we don't create new levels, we only map to existing ones)
        _logger.warning("[AI] No existing coach levels in DB — cannot map %d raw values", len(raw_values))
        _log_timing("_map_coach_levels", total_start, status="no_existing", raw=len(raw_values))
        return [], {}

    existing_info = [
        {"code": l.get("code"), "label": l.get("label")}
        for l in existing_levels
    ]
    existing_code_set = {l.get("code") for l in existing_levels}

    prompt = f"""You are a data import assistant for a padel coaching app.

I need to map level/grade/tier values found in an Excel file to the coach's
EXISTING levels in the database. The Excel values may be in any language and
may be abbreviations, full names, or codes.

Excel level values found: {json.dumps(sorted(raw_values))}

Coach's existing levels in the database (these are the ONLY valid targets):
{json.dumps(existing_info, indent=2)}

RULES:
- Each Excel value must map to EXACTLY ONE existing DB level code, or null.
- Map by meaning/similarity, not exact string match. E.g. "Iniciacao" -> "INI",
  "Avancado" -> "ADV", "Intermediate" -> "INT", etc.
- If an Excel value clearly doesn't correspond to any existing level, map it to null.
- Multiple Excel values CAN map to the same DB level code.
- You MUST only use level codes from the existing DB list above.

Return a JSON object:
{{"mapping": {{"<excel_value>": "<db_level_code or null>"}}}}"""

    content = _call_openai(
        messages=[{"role": "user", "content": prompt}],
        max_tokens=1024, use_json_mode=True,
        step_label="_map_coach_levels",
    )
    result = json.loads(content)
    raw_mapping: dict[str, str | None] = result.get("mapping", {})

    # Strictly filter: only keep mappings to actual existing DB codes
    level_mapping: dict[str, str] = {}
    for raw_val, db_code in raw_mapping.items():
        if db_code and db_code in existing_code_set:
            level_mapping[raw_val] = db_code

    # Build the output levels list: only existing levels that got matched
    matched_codes = set(level_mapping.values())
    mapped_levels = [l for l in existing_levels if l.get("code") in matched_codes]

    _log_timing("_map_coach_levels", total_start,
                raw=len(raw_values), existing=len(existing_info),
                mapped=len(level_mapping), matched_levels=len(mapped_levels))
    return mapped_levels, level_mapping


def _validate_players(
    rows: list[dict],
    level_mapping: dict[str, str],
) -> list[dict]:
    """Validate and clean Player rows.
    - Apply level_mapping to level_code
    - Keep rows that have at least a name or email
    """
    clean = []
    for row in rows:
        # Must have at least name or email
        if _is_empty(row.get("name")) and _is_empty(row.get("email")):
            continue
        # Map level_code through the coach level mapping
        raw_level = row.get("level_code")
        if raw_level and level_mapping:
            mapped = level_mapping.get(str(raw_level))
            row["level_code"] = mapped  # may be None if no match
        clean.append(row)
    return clean


def _validate_eval_categories(
    rows: list[dict],
    raw_evaluations: list[dict],
) -> list[dict]:
    """Validate Evaluation Categories.
    - Uses a quick LLM call to filter col_N / numeric / garbage names
    - Computes scale_min and scale_max from actual scores in raw_evaluations
    """
    # Collect candidate names
    candidates: list[str] = []
    for row in rows:
        name = row.get("name")
        if _is_empty(name):
            continue
        name_str = str(name).strip()
        if _is_numeric(name_str) or len(name_str) < 2:
            continue
        if name_str not in candidates:
            candidates.append(name_str)

    if not candidates:
        return []

    # Quick LLM call to filter real category names
    total_start = time.perf_counter()
    prompt = f"""I have a list of candidate evaluation category names extracted from a padel coaching spreadsheet.
Some are real categories (e.g. "Technique", "Tactics", "Physical", "Mental", "Serve", "Volley").
Others are data errors like "col_7", "col_11", "None", random codes, or non-category text.

Return a JSON object: {{"valid": ["name1", "name2", ...]}}
containing ONLY the names that are real evaluation/skill categories.

Candidates: {json.dumps(candidates)}"""

    try:
        content = _call_openai(
            messages=[{"role": "user", "content": prompt}],
            max_tokens=256, use_json_mode=True,
            step_label="_validate_eval_categories",
            count=len(candidates),
        )
        result = json.loads(content)
        valid_names: set[str] = set(result.get("valid", []))
        _log_timing("_validate_eval_categories.llm", total_start,
                    candidates=len(candidates), valid=len(valid_names))
    except Exception:
        _logger.exception("[AI] eval category LLM validation failed, using heuristic")
        # Fallback: filter out col_N patterns
        valid_names = {
            n for n in candidates
            if not re.match(r"^col_\d+$", n, re.IGNORECASE)
        }

    if not valid_names:
        return []

    # Compute scale_min and scale_max from actual evaluation scores per category
    score_ranges: dict[str, tuple[float, float]] = {}
    for ev in raw_evaluations:
        cat = ev.get("category_name")
        score = ev.get("score")
        if _is_empty(cat) or not _is_numeric(score):
            continue
        cat_str = str(cat).strip()
        s = float(score)
        if cat_str in score_ranges:
            lo, hi = score_ranges[cat_str]
            score_ranges[cat_str] = (min(lo, s), max(hi, s))
        else:
            score_ranges[cat_str] = (s, s)

    clean = []
    for name in valid_names:
        lo, hi = score_ranges.get(name, (0, 10))
        clean.append({
            "name": name,
            "scale_min": lo,
            "scale_max": hi,
        })
    return clean


def _validate_classes(rows: list[dict]) -> list[dict]:
    """Validate Classes.
    - day (date) is REQUIRED and must be valid format
    - start_time and end_time are REQUIRED and must be valid format
    - title is auto-generated from date if missing
    """
    clean = []
    for row in rows:
        day = row.get("day")
        start = row.get("start_time")
        end = row.get("end_time")

        # Date is required
        if _is_empty(day):
            continue
        day_str = str(day).strip()
        # Try to handle datetime objects
        if hasattr(day, "strftime"):
            day_str = day.strftime("%Y-%m-%d")
        # Must match YYYY-MM-DD
        if not _DATE_PATTERN.match(day_str):
            # Try to extract date from datetime string like "2025-03-15 00:00:00"
            match = re.match(r"(\d{4}-\d{2}-\d{2})", day_str)
            if match:
                day_str = match.group(1)
            else:
                continue
        row["day"] = day_str

        # Times are required
        if _is_empty(start) or _is_empty(end):
            continue
        start_str = str(start).strip()
        end_str = str(end).strip()
        # Handle time objects
        if hasattr(start, "strftime"):
            start_str = start.strftime("%H:%M")
        if hasattr(end, "strftime"):
            end_str = end.strftime("%H:%M")
        # Extract HH:MM from longer strings
        for time_str, field in [(start_str, "start_time"), (end_str, "end_time")]:
            m = re.match(r"(\d{1,2}:\d{2})", time_str)
            if m:
                row[field] = m.group(1).zfill(5)  # ensure HH:MM
            else:
                row[field] = None

        if not row.get("start_time") or not row.get("end_time"):
            continue

        # Auto-generate title if missing
        if _is_empty(row.get("title")):
            row["title"] = f"Class {day_str} {row['start_time']}-{row['end_time']}"

        clean.append(row)
    return clean


def _validate_presences(
    rows: list[dict],
    known_players: set[str],
) -> tuple[list[dict], int]:
    """Validate Presences.
    - Must have player_name in known_players
    - Must have status (present/absent)
    - Must have date
    Returns (clean_rows, dropped_count).
    """
    clean = []
    dropped = 0
    for row in rows:
        player = row.get("player_name")
        if _is_empty(player) or str(player).strip().lower() not in known_players:
            dropped += 1
            continue
        # Must have status
        status = row.get("status")
        if _is_empty(status):
            dropped += 1
            continue
        status_lower = str(status).strip().lower()
        if status_lower not in ("present", "absent"):
            dropped += 1
            continue
        row["status"] = status_lower
        # Must have date
        date_val = row.get("date")
        if _is_empty(date_val):
            dropped += 1
            continue
        date_str = str(date_val).strip()
        if hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%Y-%m-%d")
        match = re.match(r"(\d{4}-\d{2}-\d{2})", date_str)
        if not match:
            dropped += 1
            continue
        row["date"] = match.group(1)
        # Normalise justification if present
        just = row.get("justification")
        if not _is_empty(just):
            just_lower = str(just).strip().lower()
            if just_lower in ("justified", "unjustified"):
                row["justification"] = just_lower
        clean.append(row)
    return clean, dropped


def _validate_evaluations(
    rows: list[dict],
    known_players: set[str],
    known_categories: set[str],
) -> tuple[list[dict], int]:
    """Validate Evaluations.
    - Must reference known player and known category
    - Score must be numeric
    """
    clean = []
    dropped = 0
    for row in rows:
        player = row.get("player_name")
        if _is_empty(player) or str(player).strip().lower() not in known_players:
            dropped += 1
            continue
        cat = row.get("category_name")
        if _is_empty(cat) or str(cat).strip().lower() not in known_categories:
            dropped += 1
            continue
        score = row.get("score")
        if not _is_numeric(score):
            dropped += 1
            continue
        row["score"] = float(score)
        clean.append(row)
    return clean, dropped


def _validate_players_in_classes(
    rows: list[dict],
    known_players: set[str],
    known_class_titles: set[str],
) -> tuple[list[dict], int]:
    """Validate Players in Classes.
    - Must reference known player and known class title
    """
    clean = []
    dropped = 0
    for row in rows:
        player = row.get("player_name")
        title = row.get("lesson_title")
        if _is_empty(player) or str(player).strip().lower() not in known_players:
            dropped += 1
            continue
        if _is_empty(title) or str(title).strip().lower() not in known_class_titles:
            dropped += 1
            continue
        clean.append(row)
    return clean, dropped


def _has_meaningful_text(val) -> bool:
    """Check if a value contains actual meaningful text (not just symbols,
    dashes, whitespace, or empty strings)."""
    if _is_empty(val):
        return False
    text = str(val).strip()
    # Strip common placeholder characters
    cleaned = re.sub(r"[\s\-—–_.,;:!?/\\|*#@()[\]{}\"']+", "", text)
    # Must have at least 2 alphanumeric characters remaining
    return len(cleaned) >= 2


def _clean_text_value(val: str) -> str:
    """Clean up text values: strip leading bullet markers like '- ', '* ', '• '."""
    text = val.strip()
    # Strip leading list markers
    text = re.sub(r"^[-*•·>]\s+", "", text)
    # Strip trailing dangling separators
    text = text.strip(" ;,")
    return text


def _validate_player_linked(
    rows: list[dict],
    known_players: set[str],
    name_field: str = "player_name",
    text_field: str | None = None,
) -> tuple[list[dict], int]:
    """Validate tables that need a known player reference (Strengths, Weaknesses).
    If text_field is specified, also requires that field to contain meaningful text
    (not empty, not just symbols/dashes), and cleans up common formatting issues."""
    clean = []
    dropped = 0
    for row in rows:
        player = row.get(name_field)
        if _is_empty(player) or str(player).strip().lower() not in known_players:
            dropped += 1
            continue
        # If a text field is specified, require meaningful content
        if text_field:
            if not _has_meaningful_text(row.get(text_field)):
                dropped += 1
                continue
            # Clean up the text value
            row[text_field] = _clean_text_value(str(row[text_field]))
        clean.append(row)
    return clean, dropped


# ---------------------------------------------------------------------------
# Step 5b -- LLM name validation
# ---------------------------------------------------------------------------

def _validate_names_with_llm(
    names: set[str],
    context: str = "padel players",
) -> set[str]:
    """Ask LLM to filter a set of strings to only real person names.
    Returns the subset that are actual names."""
    if not names:
        return set()

    total_start = time.perf_counter()
    names_list = sorted(names)

    prompt = f"""I have a list of values that should be {context} names.
Some may be data errors (numbers, codes, random text).

Return a JSON object: {{"valid_names": ["name1", "name2", ...]}}
containing ONLY the values that are real person names.

Values to check:
{json.dumps(names_list)}"""

    content = _call_openai(
        messages=[{"role": "user", "content": prompt}],
        max_tokens=2048, use_json_mode=True,
        step_label="_validate_names_with_llm",
        count=len(names_list),
    )
    result = json.loads(content)
    valid = set(result.get("valid_names", []))
    _log_timing("_validate_names_with_llm", total_start,
                input=len(names_list), valid=len(valid))
    return valid


# ---------------------------------------------------------------------------
# Step 6 -- Cross-table player discovery
# ---------------------------------------------------------------------------

def _discover_players_from_tables(
    raw_analysis: dict[str, list[dict]],
    known_player_names: set[str],
) -> list[dict]:
    """Find player names referenced in dependent tables (Presences, Evaluations,
    Strengths, Weaknesses) that are NOT in the Players table. Ask LLM to verify
    they're real names, then create minimal Player entries for them.

    Returns new Player rows to add."""
    total_start = time.perf_counter()

    # Collect all player names from dependent tables
    referenced_names: set[str] = set()
    for table_name in ("Presences", "Evaluations", "Strengths", "Weaknesses",
                       "Players in Classes"):
        for row in raw_analysis.get(table_name, []):
            name = row.get("player_name")
            if not _is_empty(name):
                referenced_names.add(str(name).strip())

    # Find names not in known players
    unknown_names = {
        n for n in referenced_names
        if n.lower() not in known_player_names
    }

    if not unknown_names:
        _log_timing("_discover_players", total_start, unknown=0)
        return []

    # Ask LLM to verify these are real names
    valid_names = _validate_names_with_llm(unknown_names)

    new_players = []
    for name in valid_names:
        new_players.append({
            "name": name,
            "email": None,
            "phone": None,
            "level_code": None,
            "side": None,
        })

    _log_timing("_discover_players", total_start,
                unknown=len(unknown_names), valid=len(valid_names))
    return new_players


# ---------------------------------------------------------------------------
# Step 7 -- Main orchestration
# ---------------------------------------------------------------------------

def stream_import_analysis(
    file_bytes: bytes,
    coach_id: int,
    requested_tables: list[str] | None = None,
) -> Generator[str, None, None]:
    """Main entry point. Orchestrates the full import pipeline.

    Args:
        file_bytes: Raw Excel file content.
        coach_id: Coach ID for fetching existing levels from DB.
        requested_tables: List of table names to import, or None for all.
            Valid values: Players, Classes, Players in Classes, Presences,
            Evaluations, Strengths, Weaknesses.
            Coach Levels and Evaluation Categories are always auto-included.
    """
    def _event(payload: dict) -> str:
        return f"data: {json.dumps(payload)}\n\n"

    # Resolve requested tables
    if requested_tables is None:
        active_tables = set(ALL_IMPORTABLE_TABLES)
    else:
        active_tables = set(requested_tables) & set(ALL_IMPORTABLE_TABLES)
    # Always include dependency tables
    active_tables |= _AUTO_TABLES

    try:
        total_start = time.perf_counter()
        yield _event({"type": "phase", "phase": "uploading"})
        yield _event({"type": "thinking",
                       "text": f"Importing: {', '.join(sorted(active_tables - _AUTO_TABLES))}"})

        # ---- Step 1: Parse ----
        yield _event({"type": "thinking", "text": "Parsing Excel file..."})
        parse_start = time.perf_counter()
        sheets_data = parse_excel(file_bytes)
        parse_ms = _log_timing("stream.parse", parse_start, sheets=len(sheets_data))
        yield _event({"type": "thinking",
                       "text": f"Parsed in {parse_ms:.0f}ms - {len(sheets_data)} sheet(s)."})
        if not sheets_data:
            yield _event({"type": "error", "message": "No usable sheets found."})
            return

        yield _event({"type": "progress", "value": 10})

        # ---- Step 2: Relevance filter ----
        yield _event({"type": "phase", "phase": "processing"})
        yield _event({"type": "thinking", "text": "Selecting relevant sheets..."})
        rel_start = time.perf_counter()
        relevant_names = pick_relevant_sheets(sheets_data, active_tables)
        rel_ms = _log_timing("stream.relevance", rel_start, selected=len(relevant_names))
        selected = {k: v for k, v in sheets_data.items() if k in relevant_names} or sheets_data
        yield _event({"type": "thinking",
                       "text": f"Selected {len(selected)} sheet(s) in {rel_ms:.0f}ms."})
        yield _event({"type": "progress", "value": 18})

        # ---- Step 3: Segment detection ----
        yield _event({"type": "thinking", "text": "Detecting table structures..."})
        seg_start = time.perf_counter()
        all_segments: list[TableSegment] = []
        for name, raw_rows in selected.items():
            segs = detect_table_segments(name, raw_rows)
            all_segments.extend(segs)
            if len(segs) > 1:
                yield _event({"type": "thinking",
                               "text": f"Sheet '{name}': {len(segs)} separate tables detected."})
        seg_ms = _log_timing("stream.segmentation", seg_start, segments=len(all_segments))
        yield _event({"type": "thinking",
                       "text": f"{len(all_segments)} segment(s) found in {seg_ms:.0f}ms."})
        yield _event({"type": "progress", "value": 22})

        # ---- Step 4: Column mapping (parallel) ----
        yield _event({"type": "phase", "phase": "analyzing"})
        yield _event({"type": "thinking",
                       "text": f"Mapping {len(all_segments)} segment(s) via AI..."})

        from concurrent.futures import ThreadPoolExecutor, as_completed

        raw_analysis: dict[str, list[dict]] = {}
        map_start = time.perf_counter()
        with ThreadPoolExecutor(max_workers=min(len(all_segments), 4)) as pool:
            futures = {
                pool.submit(_process_segment, seg, active_tables): seg
                for seg in all_segments
            }
            done = 0
            for future in as_completed(futures):
                seg = futures[future]
                done += 1
                try:
                    seg_result = future.result()
                    _merge_into(raw_analysis, seg_result)
                    records = sum(len(v) for v in seg_result.values())
                    yield _event({"type": "thinking",
                                   "text": f"[{done}/{len(all_segments)}] "
                                           f"'{seg.label}' -> {records} records"})
                except Exception as exc:
                    _logger.exception("[AI] Failed to map segment '%s'", seg.label)
                    yield _event({"type": "thinking",
                                   "text": f"Could not map '{seg.label}': {exc}"})

        map_ms = _log_timing("stream.mapping", map_start, segments=len(all_segments))
        yield _event({"type": "thinking",
                       "text": f"Column mapping done in {map_ms:.0f}ms. "
                               f"Raw tables: {', '.join(f'{k}({len(v)})' for k, v in raw_analysis.items())}"})
        yield _event({"type": "progress", "value": 45})

        # ---- Step 5: Ordered validation chain ----
        yield _event({"type": "phase", "phase": "validating"})
        analysis: dict[str, list] = {}
        drop_counts: dict[str, int] = {}

        # 5a. Coach Levels
        yield _event({"type": "thinking", "text": "Mapping coach levels..."})
        try:
            from .coach_service import get_coach_levels
            existing_levels = get_coach_levels(coach_id)
        except (ImportError, Exception) as exc:
            _logger.warning("[AI] Could not fetch coach levels: %s", exc)
            existing_levels = []

        raw_levels = raw_analysis.get("Coach Levels", [])
        raw_players = raw_analysis.get("Players", [])
        mapped_levels, level_mapping = _map_coach_levels(raw_levels, raw_players, existing_levels)
        if mapped_levels:
            analysis["Coach Levels"] = mapped_levels
        yield _event({"type": "thinking",
                       "text": f"Coach Levels: {len(raw_levels)} found, "
                               f"{len(mapped_levels)} mapped, "
                               f"mapping: {level_mapping}"})
        yield _event({"type": "progress", "value": 50})

        # 5b. Players
        if "Players" in active_tables:
            yield _event({"type": "thinking", "text": "Validating players..."})
            raw_players = raw_analysis.get("Players", [])
            clean_players = _validate_players(raw_players, level_mapping)

            # Discover players referenced in other tables but missing from Players
            new_players = _discover_players_from_tables(raw_analysis, {
                str(p.get("name", "")).strip().lower() for p in clean_players
                if not _is_empty(p.get("name"))
            })
            if new_players:
                yield _event({"type": "thinking",
                               "text": f"Discovered {len(new_players)} additional player(s) "
                                       f"from dependent tables."})
                clean_players = _deduplicate_rows(clean_players, new_players)

            analysis["Players"] = clean_players
            yield _event({"type": "thinking",
                           "text": f"Players: {len(raw_players)} raw -> {len(clean_players)} valid"})

        # Build known players set for downstream validation
        known_player_names: set[str] = set()
        for p in analysis.get("Players", []):
            name = p.get("name")
            if not _is_empty(name):
                known_player_names.add(str(name).strip().lower())

        yield _event({"type": "progress", "value": 55})

        # 5c. Evaluation Categories
        if "Evaluation Categories" in active_tables:
            raw_cats = raw_analysis.get("Evaluation Categories", [])
            raw_evals_for_scale = raw_analysis.get("Evaluations", [])
            clean_cats = _validate_eval_categories(raw_cats, raw_evals_for_scale)
            if clean_cats:
                analysis["Evaluation Categories"] = clean_cats
            dropped = len(raw_cats) - len(clean_cats)
            if dropped:
                drop_counts["Evaluation Categories"] = dropped
            yield _event({"type": "thinking",
                           "text": f"Eval Categories: {len(raw_cats)} raw -> {len(clean_cats)} valid"
                                   f"{f' ({dropped} dropped: numeric/invalid names)' if dropped else ''}"})

        known_categories: set[str] = {
            str(c.get("name", "")).strip().lower()
            for c in analysis.get("Evaluation Categories", [])
            if not _is_empty(c.get("name"))
        }

        # 5d. Classes
        if "Classes" in active_tables:
            raw_classes = raw_analysis.get("Classes", [])
            clean_classes = _validate_classes(raw_classes)
            if clean_classes:
                analysis["Classes"] = clean_classes
            dropped = len(raw_classes) - len(clean_classes)
            if dropped:
                drop_counts["Classes"] = dropped
            yield _event({"type": "thinking",
                           "text": f"Classes: {len(raw_classes)} raw -> {len(clean_classes)} valid"
                                   f"{f' ({dropped} dropped: missing date/times)' if dropped else ''}"})

        known_class_titles: set[str] = {
            str(c.get("title", "")).strip().lower()
            for c in analysis.get("Classes", [])
            if not _is_empty(c.get("title"))
        }

        yield _event({"type": "progress", "value": 65})

        # 5e. Players in Classes
        if "Players in Classes" in active_tables:
            raw_pic = raw_analysis.get("Players in Classes", [])
            clean_pic, dropped = _validate_players_in_classes(
                raw_pic, known_player_names, known_class_titles)
            if clean_pic:
                analysis["Players in Classes"] = clean_pic
            if dropped:
                drop_counts["Players in Classes"] = dropped
            yield _event({"type": "thinking",
                           "text": f"Players in Classes: {len(raw_pic)} raw -> {len(clean_pic)} valid"
                                   f"{f' ({dropped} dropped)' if dropped else ''}"})

        # 5f. Presences
        if "Presences" in active_tables:
            raw_pres = raw_analysis.get("Presences", [])
            clean_pres, dropped = _validate_presences(raw_pres, known_player_names)
            if clean_pres:
                analysis["Presences"] = clean_pres
            if dropped:
                drop_counts["Presences"] = dropped
            yield _event({"type": "thinking",
                           "text": f"Presences: {len(raw_pres)} raw -> {len(clean_pres)} valid"
                                   f"{f' ({dropped} dropped: missing player/status/date)' if dropped else ''}"})

        yield _event({"type": "progress", "value": 75})

        # 5g. Evaluations
        if "Evaluations" in active_tables:
            raw_evals = raw_analysis.get("Evaluations", [])
            clean_evals, dropped = _validate_evaluations(
                raw_evals, known_player_names, known_categories)
            if clean_evals:
                analysis["Evaluations"] = clean_evals
            if dropped:
                drop_counts["Evaluations"] = dropped
            yield _event({"type": "thinking",
                           "text": f"Evaluations: {len(raw_evals)} raw -> {len(clean_evals)} valid"
                                   f"{f' ({dropped} dropped)' if dropped else ''}"})

        # 5h. Strengths
        if "Strengths" in active_tables:
            raw_str = raw_analysis.get("Strengths", [])
            clean_str, dropped = _validate_player_linked(
                raw_str, known_player_names, text_field="strengths")
            if clean_str:
                analysis["Strengths"] = clean_str
            if dropped:
                drop_counts["Strengths"] = dropped
            yield _event({"type": "thinking",
                           "text": f"Strengths: {len(raw_str)} raw -> {len(clean_str)} valid"
                                   f"{f' ({dropped} dropped)' if dropped else ''}"})

        # 5i. Weaknesses
        if "Weaknesses" in active_tables:
            raw_weak = raw_analysis.get("Weaknesses", [])
            clean_weak, dropped = _validate_player_linked(
                raw_weak, known_player_names, text_field="weaknesses")
            if clean_weak:
                analysis["Weaknesses"] = clean_weak
            if dropped:
                drop_counts["Weaknesses"] = dropped
            yield _event({"type": "thinking",
                           "text": f"Weaknesses: {len(raw_weak)} raw -> {len(clean_weak)} valid"
                                   f"{f' ({dropped} dropped)' if dropped else ''}"})

        yield _event({"type": "progress", "value": 85})

        # ---- Strip tables the user didn't request ----
        # (Coach Levels and Eval Categories were auto-included as dependencies
        # but only keep them in output if they have data)
        final_analysis = {
            k: v for k, v in analysis.items()
            if v  # non-empty
        }

        yield _event({"type": "progress", "value": 90})

        # ---- Summary ----
        total_records = sum(len(v) for v in final_analysis.values())
        total_dropped = sum(drop_counts.values())

        summary_parts = []
        for table_name, rows in final_analysis.items():
            part = f"{table_name}: {len(rows)}"
            if table_name in drop_counts:
                part += f" ({drop_counts[table_name]} dropped)"
            summary_parts.append(part)

        yield _event({"type": "thinking",
                       "text": f"Done - {total_records} records across {len(final_analysis)} tables"
                               f"{f', {total_dropped} rows dropped total' if total_dropped else ''}."})

        if drop_counts:
            yield _event({"type": "thinking",
                           "text": f"Dropped row summary: "
                                   + ", ".join(f"{k}: {v}" for k, v in drop_counts.items())})

        yield _event({"type": "tables", "tables": final_analysis})
        yield _event({"type": "done"})
        _log_timing("stream.total", total_start,
                    tables=len(final_analysis), records=total_records,
                    dropped=total_dropped)

    except Exception as exc:
        _logger.exception("[AI] stream_import_analysis failed")
        _log_timing("stream.total", total_start, status="failed")
        yield _event({"type": "error", "message": str(exc)})