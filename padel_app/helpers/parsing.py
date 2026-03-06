"""
Excel parsing, sheet relevance filtering, and table segment detection.
"""
from __future__ import annotations

import json
import time
from io import BytesIO

import openpyxl

from padel_app.helpers.llm import call_llm, log_timing, logger, parse_json

# ---------------------------------------------------------------------------
# Limits
# ---------------------------------------------------------------------------

MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024
MAX_SHEETS = 50
MAX_ROWS_PER_SHEET = 50_000
MAX_SEGMENTS = 100

# ---------------------------------------------------------------------------
# Step 1 — Excel parsing
# ---------------------------------------------------------------------------


def parse_excel(file_bytes: bytes) -> dict[str, list[list]]:
    """Parse workbook into {sheet_name: raw_rows}.

    Raises ValueError on oversized files.
    """
    if len(file_bytes) > MAX_FILE_SIZE_BYTES:
        raise ValueError(
            f"File too large ({len(file_bytes) / 1024 / 1024:.1f} MB, "
            f"max {MAX_FILE_SIZE_BYTES / 1024 / 1024:.0f} MB)."
        )

    total_start = time.perf_counter()
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    log_timing("parse_excel.load_workbook", total_start, sheets=len(wb.sheetnames))

    if len(wb.sheetnames) > MAX_SHEETS:
        wb.close()
        raise ValueError(f"Workbook has {len(wb.sheetnames)} sheets (max {MAX_SHEETS}).")

    sheets: dict[str, list[list]] = {}
    for name in wb.sheetnames:
        t = time.perf_counter()
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        if len(rows) < 2:
            log_timing("parse_excel.sheet", t, sheet=name, status="skipped", rows=len(rows))
            continue
        if len(rows) > MAX_ROWS_PER_SHEET:
            logger.warning("[AI] Sheet '%s': %d rows, truncating to %d", name, len(rows), MAX_ROWS_PER_SHEET)
            rows = rows[:MAX_ROWS_PER_SHEET]

        has_data = any(any(c is not None and c != "" for c in row) for row in rows[1:])
        if has_data:
            sheets[name] = rows
            log_timing("parse_excel.sheet", t, sheet=name, status="ok", rows=len(rows))
        else:
            log_timing("parse_excel.sheet", t, sheet=name, status="empty")

    wb.close()
    log_timing("parse_excel.total", total_start, sheets=len(sheets))
    return sheets


# ---------------------------------------------------------------------------
# Step 2 — Sheet relevance filter
# ---------------------------------------------------------------------------


def pick_relevant_sheets(
    sheets_data: dict[str, list[list]],
    tables_desc: str,
) -> list[str]:
    """Return sheet names worth processing.

    <=20 sheets → all. >20 → LLM filters (safety floor 33%).
    """
    if not sheets_data:
        return []
    if len(sheets_data) <= 20:
        return list(sheets_data.keys())

    total_start = time.perf_counter()
    summaries = []
    for name, rows in sheets_data.items():
        header = [str(h)[:60] if h is not None else "" for h in rows[0]]
        sample = [[str(c)[:60] if c is not None else None for c in row] for row in rows[1:4]]
        summaries.append({"sheet": name, "headers": header, "sample_rows": sample, "total_rows": len(rows) - 1})

    prompt = (
        "You are a data import assistant for a padel coaching app.\n\n"
        f"The user wants to import: {tables_desc}.\n\n"
        "IMPORTANT: Sheet names and headers may be in ANY language. "
        "Judge by the column structure and data content.\n\n"
        "Some tables require looking at sheets that aren't obviously named:\n"
        "- Evaluations, Strengths, and Weaknesses are often found in sheets "
        "about student analysis, player assessment, or evaluation.\n"
        "- Players may appear in roster, student, or member lists.\n"
        "- Presences/attendance may be in scheduling or class-related sheets.\n\n"
        "RULE: When in doubt, INCLUDE the sheet.\n\n"
        'Return a JSON object: {"sheets": ["name1", "name2", ...]}\n\n'
        f"Sheets:\n{json.dumps(summaries, default=str, indent=2)}"
    )

    try:
        content = call_llm(messages=[{"role": "user", "content": prompt}], max_tokens=512, json_mode=True, label="pick_sheets")
        result = parse_json(content, "pick_sheets")
        selected = result.get("sheets", list(sheets_data.keys()))
        if not isinstance(selected, list) or len(selected) < len(sheets_data) * 0.33:
            logger.warning("[AI] Sheet filter too aggressive, including all")
            selected = list(sheets_data.keys())
        log_timing("pick_relevant_sheets", total_start, selected=len(selected), total=len(sheets_data))
        return selected
    except Exception:
        logger.exception("[AI] pick_relevant_sheets failed, returning all")
        return list(sheets_data.keys())


# ---------------------------------------------------------------------------
# Step 3 — Table segment detection
# ---------------------------------------------------------------------------


class TableSegment:
    """A contiguous block of rows within a sheet that forms one logical table."""

    __slots__ = ("sheet_name", "segment_index", "headers", "data_rows", "start_row")

    def __init__(self, sheet_name: str, segment_index: int, headers: list[str],
                 data_rows: list[dict], start_row: int):
        self.sheet_name = sheet_name
        self.segment_index = segment_index
        self.headers = headers
        self.data_rows = data_rows
        self.start_row = start_row

    @property
    def label(self) -> str:
        return self.sheet_name if self.segment_index == 0 else f"{self.sheet_name} (segment {self.segment_index + 1})"

    def __repr__(self) -> str:
        return f"<Segment {self.label}: {len(self.data_rows)} rows, {len(self.headers)} cols>"


def _is_likely_header(row: list, prev_types: list[type] | None) -> bool:
    if not row:
        return False
    non_null = [c for c in row if c is not None and c != ""]
    if len(non_null) < 2 or not all(isinstance(c, str) for c in non_null):
        return False
    if prev_types:
        numeric = sum(1 for t in prev_types if t in (int, float))
        return numeric >= len(prev_types) * 0.3
    return False


def detect_table_segments(sheet_name: str, raw_rows: list[list]) -> list[TableSegment]:
    """Find one or more logical tables within a sheet."""
    total_start = time.perf_counter()
    if not raw_rows:
        return []

    header_idx = 0
    for i, row in enumerate(raw_rows):
        if any(c is not None and c != "" for c in row):
            header_idx = i
            break

    segments: list[TableSegment] = []
    cur_header = header_idx
    cur_data_start = header_idx + 1
    consec_blank = 0
    prev_types: list[type] | None = None

    def _finalise(data_end: int) -> None:
        nonlocal cur_header
        hdr = raw_rows[cur_header]
        headers = [str(h).strip() if h is not None else f"col_{j}" for j, h in enumerate(hdr)]
        data_rows = []
        for r in raw_rows[cur_data_start:data_end]:
            rd = {headers[j]: (r[j] if j < len(r) else None) for j in range(len(headers))}
            if not all(v is None or v == "" for v in rd.values()):
                data_rows.append(rd)
        if data_rows:
            segments.append(TableSegment(sheet_name, len(segments), headers, data_rows, cur_header))

    i = cur_data_start
    while i < len(raw_rows):
        row = raw_rows[i]
        is_blank = all(c is None or c == "" for c in row)
        if is_blank:
            consec_blank += 1
            if consec_blank >= 2:
                _finalise(i - consec_blank + 1)
                for j in range(i + 1, len(raw_rows)):
                    if any(c is not None and c != "" for c in raw_rows[j]):
                        cur_header, cur_data_start, i = j, j + 1, j + 1
                        consec_blank, prev_types = 0, None
                        break
                else:
                    i = len(raw_rows)
                continue
            i += 1
            continue
        consec_blank = 0
        if prev_types and _is_likely_header(row, prev_types):
            _finalise(i)
            cur_header, cur_data_start = i, i + 1
            prev_types = None
            i += 1
            continue
        prev_types = [type(c) for c in row]
        i += 1

    _finalise(len(raw_rows))

    if len(segments) > MAX_SEGMENTS:
        logger.warning("[AI] Sheet '%s': %d segments, capping at %d", sheet_name, len(segments), MAX_SEGMENTS)
        segments = segments[:MAX_SEGMENTS]

    log_timing("detect_segments", total_start, sheet=sheet_name, segments=len(segments))
    return segments